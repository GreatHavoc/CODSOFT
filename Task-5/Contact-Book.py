import tkinter as t
from tkinter import messagebox
import ttkbootstrap as tb
from ttkbootstrap.scrolled import ScrolledFrame
from openpyxl import load_workbook
import sqlite3

#connection to database
con=sqlite3.connect('contact_database.db')
cur=con.cursor()

#loading country codes from excel file
wb=load_workbook('country.xlsx')
ws=wb['Sheet1']
cA=ws['A']
cB=ws['B']
countrylist=[]
for i in range(len(cA)):
    countrylist.append((cA[i].value,cB[i].value))
mail_list=["outlook.com","gmail.com","ymail.com","mail.com"]

#To find the name in the contact listbox
def scankey(event):
    val=event.widget.get()
    if(len(val)==0):
        lists=[]
        lists=data
    if len(val)>0:
       lists=[]
       for item in data:
            fn,ln=item[1],item[2]
            items=fn+" "+ln
            if val.lower() in items.lower():
                lists.append(item)
    Update(lists)	 

def Update(values):
    lis.delete(0,'end')
    for i in values:
        lis.insert('end',i[1] +' '+i[2])
    lis.bind('<Double-1>',show_details)

def validate(n):
    global selected
    if(n==0):
        try:
            if(block1.get()==''):
                cur.execute("""INSERT INTO ContactDB('First_Name','Last_Name','Mobile_No','Mobile_id') VALUES (?,?,?,?)""",(fname.get(),lname.get(),mobile.get(),phoneid.get()))
                con.commit()
                messagebox.showinfo('Success','Contact Added Successfully\nPlease Restart To Take Effect')
                cms.destroy()
            else:
                cur.execute("""INSERT INTO ContactDB('First_Name','Last_Name','Mobile_No','Mobile_id','Email_Address','Email_Domain') VALUES (?,?,?,?,?,?)""",(fname.get(),lname.get(),mobile.get(),phoneid.get(),block1.get(),domain.get()))
                con.commit()
                messagebox.showinfo('Success','Contact Added Successfully\nPlease Restart To Take Effect')
                cms.destroy()
        except(sqlite3.IntegrityError):
            messagebox.showerror('Integrity Error','Number Entered by You is Too Short or Too Long or Already Exist')
            mobile.delete(0,'end')
    if(n==1):
        simple=selected[0]
        try:
            if(block1u.get()==''):
                cur.execute("""UPDATE ContactDB SET First_Name= ?,Last_Name= ?,Mobile_No= ?,Mobile_id= ? WHERE id= ? """,(fnameu.get(),lnameu.get(),mobileu.get(),phoneidu.get(),str(data[simple][0])))                
                con.commit()
                messagebox.showinfo('Success','Contact Updated Successfully')
            else:
                cur.execute("UPDATE ContactDB SET First_Name= ?,Last_Name= ?,Mobile_No= ?,Mobile_id= ?,Email_Address= ?,Email_Domain= ? WHERE id= ? """,(fnameu.get(),lnameu.get(),mobileu.get(),phoneidu.get(),block1u.get(),domainu.get(),str(data[simple][0])))
                con.commit() 
                messagebox.showinfo('Success','Contact Updated Successfully')
        except(sqlite3.IntegrityError):
            messagebox.showerror('Integrity Error','Number Entered by You is Too Short or Too Long or Already Exist')
            mobile.delete(0,'end')
    

def show_details(event):
    global selected
    selected=lis.curselection()
    simple=selected[0]

    person_details=tb.Toplevel(title='Person Details',size=(350,280),position=(1100,300),resizable=(False,False))
    person_details.iconbitmap('contact.ico')
    person_details.after(10000,person_details.destroy)
    tb.Label(person_details,font=('Helevatica',13),style='primary',text=data[simple][0]).place(x=160,y=2)
    tb.Label(person_details,bootstyle='primary',font=('Georgia',12),text='Person Id').place(x=128,y=30)

    tb.Separator(person_details,bootstyle='info').place(x=0,y=57,width=350)
    tb.LabelFrame(person_details,text="First Name:",style='primary',height=60,width=160).grid(row=0,column=0,sticky='w',pady=(60,5),padx=(5,10))
    tb.Label(person_details,text=data[simple][1],font=('Georgia',12),style='info').place(y=80,x=15)
    tb.LabelFrame(person_details,text="Last Name:",style='primary',height=60,width=160).grid(row=0,column=1,sticky='w',pady=(60,5),padx=(10,5))
    tb.Label(person_details,text=data[simple][2],font=('Georgia',12),style='info').place(y=80,x=200)
    tb.LabelFrame(person_details,bootstyle='primary',text='ISD code',height=60,width=113).place(x=8,y=130)
    tb.Label(person_details,text=data[simple][4],font=('Helevatica',12),style='info').place(x=12,y=150)
    tb.LabelFrame(person_details,bootstyle='primary',text='Mobile No.',height=60,width=218).place(x=127,y=130)
    tb.Label(person_details,text=data[simple][3],font=('Helevatica',12),style=('info')).place(x=175,y=150)
    tb.LabelFrame(person_details,text="Email Address",style='primary',height=60,width=337).grid(row=2,columnspan=2,pady=75)
    tb.Label(person_details,text=data[simple][5]+'@'+data[simple][6],font=('Georgia',10),style='info').place(x=12,y=220)


 
def update_window():
    global selected
    selected=lis.curselection()
    if(selected!=()):
        simple=int(str(selected)[1])
        ums=tb.Toplevel(title='Update Contact Page',size=(400,470),position=(780,250))
        ums.iconbitmap('contact.ico')

        global fnameu,lnameu,phoneidu,mobileu,block1u,domainu,block1u
        
        tb.Label(ums,bootstyle='primary',font=('Georgia',12),text='Enter the contact details below: ').grid(row=0,padx=(6,0),pady=(10,0))

        tb.Separator(ums,bootstyle='info-dotted').place(x=0,y=50,width=400)

        nameu=tb.LabelFrame(ums,bootstyle='info',text='Name: ',height=100,width=390)
        nameu.grid(row=2,column=0,padx=(8,0),pady=(20,5),sticky='n,e,w,s')

        tb.Label(nameu,text="First Name:",font=('Georgia',12),style='primary').grid(row=0,column=0,pady=(7,5),padx=10)

        fnameu=tb.Entry(nameu,style='info',font=("Gerogia",12),justify='center')
        fnameu.grid(row=0,column=1)
        fnameu.insert(0,data[simple][1])
        tb.Label(nameu,text="Last Name:",font=('Georgia',12),style='primary').grid(row=1,column=0,pady=10,padx=10)
        lnameu=tb.Entry(nameu,style='info',font=("Gerogia",12),justify='center')
        lnameu.grid(row=1,column=1)
        lnameu.insert(0,data[simple][2])

        phoneu=tb.LabelFrame(ums,bootstyle='info',text='Phone No.: ',height=100,width=390)
        phoneu.grid(row=3,column=0,padx=(8,0),pady=5,sticky='n,e,w,s')
        tb.Label(phoneu,text="Select Country:",font=('Georgia',12),style='primary').grid(row=0,column=0,pady=10,padx=10,sticky='w')

        phoneidu=tb.Combobox(phoneu,style='info',values=countrylist)
        phoneidu.grid(row=0,column=1,sticky='w',padx=10,pady=10)
        phoneidu.set(data[simple][4])

        tb.Label(phoneu,text="Mobile No.:",font=('Georgia',11),style='primary').grid(row=1,column=0,pady=10,padx=10,sticky='w')
        mobileu=tb.Entry(phoneu,style='info',font=("Gerogia",11),width=18)
        mobileu.grid(row=1,column=1,padx=10,pady=10)
        mobileu.insert(0,data[simple][3])

        emailu=tb.LabelFrame(ums,bootstyle='info',text='Email: ',height=100,width=390)
        emailu.grid(row=4,column=0,padx=(8,0),pady=5,sticky='n,e,w,s')
        block1u=tb.Entry(emailu,style='primary',font=('Georgia',11),width=18)
        block1u.grid(row=0,column=0,padx=(5,3),pady=10,sticky='w')
        block1u.insert(0,data[simple][5])

        tb.Label(emailu,text='@',font=('Georgia',11),style='primary').grid(row=0,column=1,padx=3,pady=(0,7))
        domainu=tb.Combobox(emailu,style='info',values=mail_list,width=11)
        domainu.grid(row=0,column=2,padx=4,pady=10)
        domainu.set(data[simple][6])

        update=tb.Button(ums,style='danger,outline',text='            +\nUpdate Contact',command=lambda:validate(1))
        update.grid(row=5,padx=50,pady=7)


        ums.mainloop()
    else:
        messagebox.showerror('Selection Error','Please select one Contact from the list by clicking it.')

    

def create_window():

    global fname,lname,phoneid,mobile,block1,domain,block1,cms

    cms=tb.Toplevel(title='Create Contact Page',size=(400,470),position=(980,250))
    cms.iconbitmap('contact.ico')

    info=tb.Label(cms,bootstyle='primary',font=('Georgia',12),text='Enter the contact details below: ').grid(row=0,padx=(6,0),pady=(10,0))

    sep1=tb.Separator(cms,bootstyle='info-dotted').place(x=0,y=50,width=400)

    name=tb.LabelFrame(cms,bootstyle='info',text='Name: ',height=100,width=390)
    name.grid(row=2,column=0,padx=(8,0),pady=(20,5),sticky='n,e,w,s')

    tb.Label(name,text="First Name:",font=('Georgia',12),style='primary').grid(row=0,column=0,pady=(7,5),padx=10)

    fname=tb.Entry(name,style='info',font=("Gerogia",12))
    fname.grid(row=0,column=1)
    tb.Label(name,text="Last Name:",font=('Georgia',12),style='primary').grid(row=1,column=0,pady=10,padx=10)
    lname=tb.Entry(name,style='info',font=("Gerogia",12))
    lname.grid(row=1,column=1)

    phone=tb.LabelFrame(cms,bootstyle='info',text='Phone No.: ',height=100,width=390)
    phone.grid(row=3,column=0,padx=(8,0),pady=5,sticky='n,e,w,s')
    tb.Label(phone,text="Select Country:",font=('Georgia',12),style='primary').grid(row=0,column=0,pady=10,padx=10,sticky='w')

    phoneid=tb.Combobox(phone,style='info',values=countrylist)
    phoneid.grid(row=0,column=1,sticky='w',padx=10,pady=10)
    phoneid.current(94)

    tb.Label(phone,text="Mobile No.:",font=('Georgia',11),style='primary').grid(row=1,column=0,pady=10,padx=10,sticky='w')
    mobile=tb.Entry(phone,style='info',font=("Gerogia",11),width=18)
    mobile.grid(row=1,column=1,padx=10,pady=10)

    email=tb.LabelFrame(cms,bootstyle='info',text='Email: ',height=100,width=390)
    email.grid(row=4,column=0,padx=(8,0),pady=5,sticky='n,e,w,s')
    block1=tb.Entry(email,style='primary',font=('Georgia',11),width=18)
    block1.grid(row=0,column=0,padx=(5,3),pady=10,sticky='w')
    tb.Label(email,text='@',font=('Georgia',11),style='primary').grid(row=0,column=1,padx=3,pady=(0,7))
    domain=tb.Combobox(email,style='info',values=mail_list,width=11)
    domain.grid(row=0,column=2,padx=4,pady=10)
    domain.current(1)

    add=tb.Button(cms,style='danger,outline',text='         +\nAdd Contact',command=lambda:validate(0))
    add.grid(row=5,padx=50,pady=7)

    cms.mainloop()

def delete_contact():
    selected=lis.curselection()
    if(selected!=()):
        simple=selected[0]
        cur.execute("DELETE FROM ContactDB WHERE id= {} ".format(data[simple][0]))
        con.commit()
        del data[simple]
        lis.delete(simple)
        messagebox.showinfo('Success','Contact Deleted Successfully')
    else:
        messagebox.showerror('Selection Error','Please Select The Contact To Delete')
    return
root=tb.Window(themename='superhero',title='Contacts App',size=(500,600),position=(500,250),resizable=(0,0))
root.iconbitmap('contact.ico')

create=tb.Button(root,text='         + \nCreate New',style='success outline',command=create_window,width=12)
update=tb.Button(root,text='          +- \nUpdate Existing',style='info outline',command=update_window)
delete=tb.Button(root,text='          - \nDelete Contact',style='danger outline',command=delete_contact)
create.place(y=30,x=25)
update.place(y=30,x=184)
delete.place(y=30,x=350)

sep=tb.Separator(root,bootstyle='info dotted')
sep.place(y=96,x=0,relwidth=1,width=500)
sep=tb.Separator(root,bootstyle='info dotted')
sep.place(y=170,x=0,relwidth=1,width=500)

label=tb.Label(root,text='Search Box->',style='secondary.TLabel',font=('Georgia',15)).place(x=10,y=117)

search_box=tb.Entry(root,style='info.TEntry')
search_box.place(x=280,y=118)
search_box.bind('<KeyRelease>',scankey)

sf=ScrolledFrame(root, autohide=True,bootstyle='dark rounded',height=420,width=410)
sf.pack(pady=(175,4),expand=t.NO,padx=5)

lis=t.Listbox(sf,height=14,font=('Georgia',14),justify='center',highlightcolor='blue')
lis.pack(expand=t.NO,fill=t.BOTH)
cur.execute('SELECT * FROM ContactDB')
data=cur.fetchall()

Update(data)


root.mainloop()